import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

from backend.list_nums import list_nums
from backend.generate_starting_values import generate_starting_values
from backend.verify_instructions import verify_instructions
from backend.run_vb_script import run_vb_script


def Create_Dashboard(metric1, metric2, metric3, metric4 ,metric5):
    print(f"Metrics entered: {metric1}, {metric2}, {metric3}, {metric4}, {metric5}")
    while True:
        metric_count = 0
        instructions = {}
        dashboard_height = int(metric1)
        dashboard_width = int(metric2)
        num_of_metrics = int(metric3)
        min_height = int(metric4)
        min_width = int(metric5)
        #Generates Number of Rows, and Number of Metrics Displayed in a Row
        num_of_metrics_in_a_row = list_nums(num_of_metrics,num_of_metrics)
        num_of_metrics_in_a_row = [num for num in num_of_metrics_in_a_row if num !=0]
        #Generates row heights row heights must be greater than a specified value
        row_heights = list_nums(dashboard_height,len(num_of_metrics_in_a_row))
        while any (value <= min_height for value in row_heights):
                row_heights = list_nums(dashboard_height,len(num_of_metrics_in_a_row))
        #Generates row heights with starting cordinates for instance a row height of ten is now a list 1-10 meanning the row consists of rows 1 through 10            
        row_heights_with_starting_coordinates = generate_starting_values(row_heights)

        #Generates the Lengths of Each metric that exists in a row. If a metric has a width that is below an entered value than the a new set of widths is generated
        for count, row in enumerate(num_of_metrics_in_a_row):
            shape_lengths = list_nums(dashboard_width, row)
            while any (value <= min_width for value in  shape_lengths):
                shape_lengths = list_nums(dashboard_width, row)
            instructions[count] =(row_heights_with_starting_coordinates[count], shape_lengths)


        #Verifys the dashboard instructions if the instructions are valid, then a new dashboard is generated
        validity = verify_instructions(instructions)
        print("Loading...")
        if validity == True:
            print('Success!') 
            break

#builds dashboard
    workbook = Workbook()

    #Select the active sheet
    sheet = workbook.active


    ##Color Entire Range
    for row in range(1,dashboard_height):
            for col in range(1,dashboard_width): 
                color = "B2B2B2"
                cell = sheet.cell(row=row, column=col)
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                sheet.row_dimensions[row].height = 15
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 3


    #Carve dashboards based on dimensions
    for num in range(0,len(num_of_metrics_in_a_row)):
        print("Row",num,"\nContains: ", len(instructions[num][1]),"Metrics","\n Metric Lengths ",instructions[num][1], "\n---------------------------")
        
        #Trace Row
        for col in range(1,dashboard_width):
            row = instructions[num][0][1]
            color = "FFFFFF"
            cell = sheet.cell(row=row, column=col)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid") 
            sheet.row_dimensions[row+1].height = 5

        #Trace col
        for num2 in range(1,len(instructions[num][1])):
            for row in range(instructions[num][0][0],instructions[num][0][1]):
                color = "FFFFFF"
                col = instructions[num][1][num2]
                cell = sheet.cell(row=row, column=col)
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col+1)].width = 1
        ##Add Titles
        titles = []
        instructions[num][1].pop(0)
        titles.append(1)
        for col in reversed(instructions[num][1]):
            titles.append(col+1)
        for col in titles:
            metric_count = metric_count + 1
            row = instructions[num][0][0]
            cell = sheet.cell(row=row, column=col)

            font_style = Font(name='Arial', size=12, bold=True, color='000000')
            cell.font = font_style
            cell.alignment = Alignment(horizontal='left',vertical='top')
            cell.value = "Metric_"+ str(metric_count)
            new_sheet = workbook.create_sheet(title="dataMetric_"+ str(metric_count))

    ##Optimally Resize Columns
    intialiter = True
    for col in range(2,dashboard_width):
        if intialiter == True:
            firstcol = col
            count = 0
            intialiter = False
        if sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width <= 1:
            sheet.column_dimensions[openpyxl.utils.get_column_letter(firstcol)].width = count * 3
            count = 0
            intialiter = True
        if col == dashboard_width-1:
            firstcol = dashboard_width - count - 1
            sheet.column_dimensions[openpyxl.utils.get_column_letter(firstcol)].width = count * 3
        count = count + 1
    ##Optimally Resize Rows
    intialiter = True
    for row in range(2,dashboard_height):
        if intialiter == True:
            firstrow = row
            count = 0
            intialiter = False
        if sheet.row_dimensions[row].height < 15:
            sheet.row_dimensions[firstrow].height = count * 15
            count = 0
            intialiter = True
        if row == dashboard_height-1:
            firstrow = dashboard_height - count - 1
            sheet.row_dimensions[firstrow].height = count * 15
        count = count + 1
    sheet.insert_cols(1)
    sheet.column_dimensions['A'].width = 15   
    sheet.title = "Dashboard"
    sheet.insert_rows(1)


    #Label which cols and rows to remove
    for col in range(2,dashboard_width+1):
        cell = sheet.cell(row=1, column=col)
        cell.value = sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width   

    for row in range(2,dashboard_height+1):
        cell = sheet.cell(row=row, column=1)
        cell.value = sheet.row_dimensions[row].height

    sheet.cell(row=dashboard_height, column=1).value = 15
    sheet.cell(row=1, column=dashboard_width).value = 3

    workbook.save("Dashboard_.xlsx")
    #custom VB script that deletes rows, and cols
    run_vb_script()
